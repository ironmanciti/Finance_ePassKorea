# %% [markdown]
# 37차시: AI 분석 리포트 자동화 (PDF/Excel)
# =========================================
#
# 학습 목표:
# - Python으로 PDF 리포트 자동 생성
# - Excel 보고서에 데이터와 차트 삽입
# - 분석 결과를 문서로 정리하는 파이프라인 구축
#
# 학습 내용:
# 1. PDF 리포트 생성 (reportlab)
# 2. Excel 보고서 생성 (openpyxl)
# 3. 차트 이미지 생성
# 4. 통계 계산 함수
# 5. PDF 리포트 생성 함수 (독립 실행)
# 6. Excel 리포트 생성 함수 (독립 실행)

# %%
# # !pip install -Uq reportlab koreanize-matplotlib

# %%
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import koreanize_matplotlib
import pandas as pd
import numpy as np
import os
from datetime import datetime, date, timedelta
from pykrx import stock


# %% [markdown]
# ## 1. 폰트 설정
#
# 한글 폰트를 등록하여 PDF와 차트에서 한글이 정상적으로 표시되도록 합니다.

# %%
# ============================================
# 1. 한글 폰트 등록
# ============================================
font_paths = [
    'C:/Windows/Fonts/malgun.ttf',  # Windows
    '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',  # Linux/Colab
    '/System/Library/Fonts/AppleGothic.ttf'  # Mac
]

font_registered = False
for path in font_paths:
    # 해당 경로에 폰트 파일이 실제로 존재하는지 확인
    if os.path.exists(path):
        try:
            # 폰트 파일을 ReportLab에 'Korean' 이라는 이름으로 등록
            pdfmetrics.registerFont(TTFont('Korean', path))
            print(f"[폰트 등록 성공]: {path}")
            # 한글 폰트 등록 성공 여부 플래그 설정
            font_registered = True
            break
        except Exception as e:
            print(f"[폰트 등록 실패]: {e}")
            continue

if not font_registered:
    print("[경고] 한글 폰트를 찾을 수 없습니다. 영문만 사용됩니다.")

# ============================================
# 출력 폴더 설정
# ============================================
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# %% [markdown]
# ## 2. PDF 리포트 생성
#
# reportlab을 사용하여 PDF 리포트를 생성합니다.

# %% [markdown]
# ### 2.1 PDF 기본 생성
#
# 간단한 텍스트만 포함된 PDF를 생성합니다.

# %%
# --------------------------------------------
# 2.1 간단한 PDF 생성 예제
# --------------------------------------------
def create_simple_pdf(filename: str):
    
    # PDF 문서 객체 생성 (A4 용지 크기)
    doc = SimpleDocTemplate(filename, pagesize=A4)
    
    # ReportLab 기본 스타일 시트 로드
    styles = getSampleStyleSheet()
    
    if font_registered:
        # 본문용 한글 스타일 정의
        styles.add(ParagraphStyle(
            name='Korean',        # 스타일 이름
            fontName='Korean',    # 등록된 한글 폰트 이름
            fontSize=12,          # 기본 글자 크기
            leading=16            # 줄 간격 (행간)
        ))
        
        # 제목용 한글 스타일 정의
        styles.add(ParagraphStyle(
            name='KoreanTitle',   # 제목 스타일 이름
            fontName='Korean',    # 동일한 한글 폰트 사용
            fontSize=18,          # 제목 글자 크기
            leading=22,           # 제목 행간
            spaceAfter=20         # 제목 아래 여백
        ))

    story = []
    
    # 제목
    if font_registered:
        story.append(Paragraph("주식 분석 리포트", styles['KoreanTitle']))
    else:
        story.append(Paragraph("Stock Analysis Report", styles['Title']))
    
    story.append(Spacer(1, 20))
    
    # 내용
    content = "이 리포트는 Python으로 자동 생성되었습니다."
    if font_registered:
        story.append(Paragraph(content, styles['Korean']))
    else:
        story.append(Paragraph("This report was auto-generated.", styles['Normal']))
    
    # PDF 생성
    doc.build(story)
    return filename

# %% [markdown]
# ### 2.2 PDF에 테이블 추가
#
# 테이블이 포함된 PDF를 생성합니다.

# %%
# --------------------------------------------
# 2.2 PDF에 테이블 추가
# --------------------------------------------
def create_pdf_with_table(filename: str, data: list, headers: list, title: str):
    """
    테이블이 포함된 PDF 생성
    
    Parameters:
        filename: 저장할 파일명
        data: 2차원 리스트 (행 데이터)
        headers: 컬럼 헤더
        title: 리포트 제목
    """
    doc = SimpleDocTemplate(filename, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # 제목
    story.append(Paragraph(title, styles['Title']))
    story.append(Spacer(1, 20))
    
    # 테이블 데이터 (헤더 + 데이터)
    table_data = [headers] + data
    
    # 테이블 생성
    table = Table(table_data)
    
    # 테이블 스타일 설정 
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),        # 첫 번째 행 배경색 설정, (row, col)
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),   # 헤더 텍스트 색상 설정
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),               # 테이블 전체 텍스트 가운데 정렬
        ('FONTSIZE', (0, 0), (-1, -1), 10),                  # 테이블 전체 글자 크기 설정
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),              # 헤더 행 하단 여백 추가
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),      # 헤더를 제외한 본문 행 배경색
        ('GRID', (0, 0), (-1, -1), 1, colors.black)          # 테이블 전체에 격자선 추가
    ]))

    story.append(table)
    
    doc.build(story)
    return filename

# %% [markdown]
# ### 2.3 PDF에 차트 이미지 삽입
#
# 차트 이미지가 포함된 PDF를 생성합니다.

# %%
# --------------------------------------------
# 2.3 PDF에 차트 이미지 삽입
# --------------------------------------------
def create_pdf_with_chart(filename: str, title: str, content: str, chart_path: str):
    """차트 이미지가 포함된 PDF 생성"""
    doc = SimpleDocTemplate(filename, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # 제목
    story.append(Paragraph(title, styles['Title']))
    story.append(Spacer(1, 20))
    
    # 내용
    story.append(Paragraph(content, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # 차트 이미지
    if os.path.exists(chart_path):
        img = Image(chart_path, width=15*cm, height=8*cm)
        story.append(img)
    
    doc.build(story)
    return filename

# %% [markdown]
# ## 3. 차트 이미지 생성
#
# matplotlib을 사용하여 주가 차트 이미지를 생성합니다.

# %%
# ============================================
# 3. 차트 이미지 생성
# ============================================
def create_stock_chart_image(df: pd.DataFrame, stock_name: str, output_path: str):
    """주가 차트 이미지 생성"""
    fig, axes = plt.subplots(2, 1, figsize=(10, 8), gridspec_kw={'height_ratios': [3, 1]})
    
    # 종가 차트
    axes[0].plot(df.index, df['종가'], 'b-', linewidth=1.5)
    axes[0].set_title(f'{stock_name} 주가 추이')
    axes[0].set_ylabel('주가 (원)')
    axes[0].grid(True, alpha=0.3)
    
    # 거래량 차트
    bar_colors = ['red' if c >= o else 'blue' for c, o in zip(df['종가'], df['시가'])]
    axes[1].bar(df.index, df['거래량'], color=bar_colors, alpha=0.7)
    axes[1].set_ylabel('거래량')
    axes[1].set_xlabel('날짜')
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close()
    
    # 이미지 파일이 제대로 생성되었는지 확인
    if not os.path.exists(output_path):
        raise FileNotFoundError(f"차트 이미지 파일이 생성되지 않았습니다: {output_path}")
    
    print(f"[차트 이미지 생성 완료]: {output_path}")
    return output_path

# %% [markdown]
# ## 4. 통계 계산 함수
#
# 주가 데이터로부터 통계를 계산합니다.

# %%
# ============================================
# 4. 통계 계산 함수
# ============================================
def calculate_stock_stats(df: pd.DataFrame) -> dict:
    """
    주가 데이터로부터 통계를 계산합니다.
    
    Parameters:
        df: OHLCV 데이터
    
    Returns:
        dict: 통계 정보 딕셔너리
    """
    stats = {
        '시작일': df.index[0].strftime('%Y-%m-%d'),
        '종료일': df.index[-1].strftime('%Y-%m-%d'),
        '시작가': f"{df['종가'].iloc[0]:,.0f}원",
        '종료가': f"{df['종가'].iloc[-1]:,.0f}원",
        '최고가': f"{df['고가'].max():,.0f}원",
        '최저가': f"{df['저가'].min():,.0f}원",
        '평균 거래량': f"{df['거래량'].mean():,.0f}",
        '기간 수익률': f"{((df['종가'].iloc[-1] / df['종가'].iloc[0]) - 1) * 100:.2f}%"
    }
    return stats

# %% [markdown]
# ## 5. PDF 리포트 생성 함수
#
# 주식 데이터를 분석하여 PDF 리포트를 생성합니다.

# %%
# ============================================
# 5. PDF 리포트 생성 함수
# ============================================
def generate_pdf_report(stock_code: str, stock_name: str, df: pd.DataFrame, 
                       chart_path: str, output_dir: str = ".") -> str:
    """
    주식 분석 PDF 리포트 생성
    
    Parameters:
        stock_code: 종목코드
        stock_name: 종목명
        df: OHLCV 데이터
        chart_path: 차트 이미지 경로
        output_dir: 출력 디렉토리
    
    Returns:
        str: 생성된 PDF 파일 경로
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_path = os.path.join(output_dir, f"report_{stock_code}_{timestamp}.pdf")
    
    # 통계 계산
    stats = calculate_stock_stats(df)
    
    # PDF 문서 생성
    doc = SimpleDocTemplate(pdf_path, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # 한글 스타일 추가 (폰트가 등록된 경우)
    if font_registered:
        if 'Korean' not in [s.name for s in styles.byName.values()]:
            styles.add(ParagraphStyle(
                name='Korean',
                fontName='Korean',
                fontSize=12,
                leading=16
            ))
            styles.add(ParagraphStyle(
                name='KoreanTitle',
                fontName='Korean',
                fontSize=18,
                leading=22,
                spaceAfter=20
            ))
    
    story = []
    
    # 제목
    if font_registered:
        story.append(Paragraph(f"{stock_name} ({stock_code}) 분석 리포트", styles['KoreanTitle']))
    else:
        story.append(Paragraph(f"{stock_name} ({stock_code}) Analysis Report", styles['Title']))
    story.append(Spacer(1, 20))
    
    # 통계 테이블
    if font_registered:
        table_headers = [['항목', '값']]
    else:
        table_headers = [['Item', 'Value']]
    
    table_data = [[k, v] for k, v in stats.items()]
    table = Table(table_headers + table_data)
    
    # 테이블 스타일 설정
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE', (0, 0), (-1, -1), 10)
    ]
    
    if font_registered:
        table_style.append(('FONTNAME', (0, 0), (-1, -1), 'Korean'))
    
    table.setStyle(TableStyle(table_style))
    story.append(table)
    story.append(Spacer(1, 20))
    
    # 차트 이미지 삽입
    if os.path.exists(chart_path):
        try:
            img = Image(chart_path, width=15*cm, height=10*cm)
            story.append(img)
            print(f"[PDF에 차트 이미지 삽입 성공]")
        except Exception as e:
            print(f"[PDF에 차트 이미지 삽입 실패]: {e}")
            if font_registered:
                story.append(Paragraph(f"[차트 이미지 로드 실패: {str(e)}]", styles['Korean']))
            else:
                story.append(Paragraph(f"[차트 이미지 로드 실패: {str(e)}]", styles['Normal']))
    else:
        error_msg = f"[차트 이미지 파일을 찾을 수 없습니다: {chart_path}]"
        if font_registered:
            story.append(Paragraph(error_msg, styles['Korean']))
        else:
            story.append(Paragraph(error_msg, styles['Normal']))
    
    doc.build(story)
    print(f"[PDF 리포트 생성 완료]: {pdf_path}")
    return pdf_path

# %% [markdown]
# ## 6. PDF 리포트 생성 실행
#
# 실제 데이터를 수집하여 PDF 리포트를 생성합니다.

# %%
# ============================================
# 6. PDF 리포트 생성 실행
# ============================================
# 데이터 수집
end = date.today()
start = end - timedelta(days=90)

df_samsung = stock.get_market_ohlcv(
    start.strftime("%Y%m%d"),
    end.strftime("%Y%m%d"),
    "005930"
)

print(f"[데이터 수집 완료]")
print(f"  종목: 삼성전자 (005930)")
print(f"  기간: {start} ~ {end}")
print(f"  데이터 수: {len(df_samsung)}일")
print()

# 차트 이미지 생성
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
chart_path = os.path.join(OUTPUT_DIR, f"chart_005930_{timestamp}.png")
create_stock_chart_image(df_samsung, "삼성전자", chart_path)
print()

# PDF 리포트 생성
print("[PDF 리포트 생성 중...]")
pdf_path = generate_pdf_report(
    stock_code="005930",
    stock_name="삼성전자",
    df=df_samsung,
    chart_path=chart_path,
    output_dir=OUTPUT_DIR
)
print()

print("=" * 60)
print("PDF 리포트 생성 완료!")
print("=" * 60)
print(f"PDF:   {pdf_path}")
print(f"Chart: {chart_path}")

# %% [markdown]
# ## 7. Excel 보고서 생성
#
# openpyxl을 사용하여 Excel 보고서를 생성합니다.

# %% [markdown]
# ### 7.1 Excel 기본 생성
#
# DataFrame을 Excel 보고서로 변환합니다.

# %%
# --------------------------------------------
# 7.1 Excel 기본 생성
# --------------------------------------------
def create_excel_report(filename: str, df: pd.DataFrame, title: str):
    """
    DataFrame을 Excel 보고서로 변환
    
    Parameters:
        filename: 저장할 파일명
        df: pandas DataFrame
        title: 시트 제목
    """
    wb = openpyxl.Workbook()   # 새로운 Excel 워크북(파일) 생성
    ws = wb.active     # 기본으로 생성되는 활성 시트 가져오기
    ws.title = "Report"   # 시트 이름을 'Report'로 변경
    
    # 제목용 글꼴 스타일
    title_font = Font(bold=True, size=14)
    
    # 헤더(컬럼명)용 글꼴 스타일
    header_font = Font(bold=True, color="FFFFFF")
    
    # 헤더 셀 배경색 스타일
    header_fill = PatternFill(
        start_color="4472C4",   # 파란색 계열(#4472C4)
        end_color="4472C4",
        fill_type="solid"      # 단색 채우기
    )
    
    # 테이블 테두리 스타일
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 제목
    ws['A1'] = title
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    
    start_row = 3  # DataFrame 추가 (3행부터)
    
    # DataFrame을 행(row) 단위로 변환하여 Excel 시트에 기록
    # - index=False : DataFrame 인덱스 제외
    # - header=True : 컬럼명을 첫 행에 포함
    # - start_row부터 데이터 쓰기 시작
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True),
        start_row
    ):
        for c_idx, value in enumerate(row, 1):    # 각 행의 값을 열(column) 단위로 순회
            cell = ws.cell(row=r_idx, column=c_idx, value=value)   # 지정한 행/열 위치에 셀 생성 및 값 입력
            cell.border = border   # 셀 테두리 적용
            cell.alignment = Alignment(horizontal='center')   # 셀 내용 가운데 정렬

            # 헤더 행 스타일 적용
            if r_idx == start_row:
                cell.font = header_font      # 헤더 글꼴 스타일 적용
                cell.fill = header_fill      # 헤더 배경색 적용
    
    # ----------------------------------------
    # 컬럼 너비 자동 조정
    # - 각 컬럼의 내용 길이를 기준으로 폭 설정
    # ----------------------------------------
    for column in ws.columns:
        max_length = 0
        
        # 현재 컬럼의 알파벳 이름
        column_letter = column[0].column_letter
        
        # 컬럼 내 모든 셀을 순회하며 가장 긴 텍스트 길이 계산
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # 계산된 최대 길이에 여유 폭(+2)을 더해 컬럼 너비 설정
        ws.column_dimensions[column_letter].width = max_length + 2
    
    wb.save(filename)
    return filename

# %% [markdown]
# ### 7.2 Excel에 차트 이미지 삽입
#
# Excel 파일에 차트 이미지를 추가합니다.

# %%
# --------------------------------------------
# 7.3 Excel에 차트 이미지 삽입
# --------------------------------------------
def add_chart_to_excel(filename: str, chart_path: str, cell: str = 'F3'):
    wb = openpyxl.load_workbook(filename)  # 기존 Excel 파일 열기
    ws = wb.active        # 현재 활성화된 시트 가져오기
    
    if os.path.exists(chart_path):
        img = XLImage(chart_path)  # 이미지 파일을 Excel 이미지 객체로 로드
        img.width = 500    # 이미지 가로 크기 설정 (픽셀 단위)
        img.height = 300   # 이미지 세로 크기 설정 (픽셀 단위)
        
        # 지정한 셀 위치에 이미지 삽입
        ws.add_image(img, cell)

    wb.save(filename)
    return filename

# %% [markdown]
# ## 8. Excel 리포트 생성 함수
#
# 주식 데이터를 분석하여 Excel 리포트를 생성합니다.

# %%
# ============================================
# 8. Excel 리포트 생성 함수
# ============================================
def generate_excel_report(stock_code: str, stock_name: str, df: pd.DataFrame,
                          chart_path: str, output_dir: str = ".") -> str:
    """
    주식 분석 Excel 리포트 생성
    
    Parameters:
        stock_code: 종목코드
        stock_name: 종목명
        df: OHLCV 데이터
        chart_path: 차트 이미지 경로
        output_dir: 출력 디렉토리
    
    Returns:
        str: 생성된 Excel 파일 경로
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"report_{stock_code}_{timestamp}.xlsx")
    
    # 통계 계산
    stats = calculate_stock_stats(df)
    
    # Excel 파일 생성
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # 주가 데이터 시트
        df.to_excel(writer, sheet_name='주가데이터')
        
        # 통계 시트
        stats_df = pd.DataFrame(list(stats.items()), columns=['항목', '값'])
        stats_df.to_excel(writer, sheet_name='통계', index=False)
    
    # 차트 이미지 추가
    if os.path.exists(chart_path):
        add_chart_to_excel(excel_path, chart_path, 'E3')
        print(f"[Excel에 차트 이미지 추가 완료]")
    
    print(f"[Excel 리포트 생성 완료]: {excel_path}")
    return excel_path

# %% [markdown]
# ## 9. Excel 리포트 생성 실행
#
# 실제 데이터를 수집하여 Excel 리포트를 생성합니다.

# %%
# ============================================
# 9. Excel 리포트 생성 실행
# ============================================
# 데이터 수집
end = date.today()
start = end - timedelta(days=90)

df_samsung = stock.get_market_ohlcv(
    start.strftime("%Y%m%d"),
    end.strftime("%Y%m%d"),
    "005930"
)

print(f"[데이터 수집 완료]")
print(f"  종목: 삼성전자 (005930)")
print(f"  기간: {start} ~ {end}")
print(f"  데이터 수: {len(df_samsung)}일")
print()

# 차트 이미지 생성
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
chart_path = os.path.join(OUTPUT_DIR, f"chart_005930_{timestamp}.png")
create_stock_chart_image(df_samsung, "삼성전자", chart_path)
print()

# Excel 리포트 생성
print("[Excel 리포트 생성 중...]")
excel_path = generate_excel_report(
    stock_code="005930",
    stock_name="삼성전자",
    df=df_samsung,
    chart_path=chart_path,
    output_dir=OUTPUT_DIR
)
print()

print("=" * 60)
print("Excel 리포트 생성 완료!")
print("=" * 60)
print(f"Excel: {excel_path}")
print(f"Chart: {chart_path}")

# %% [markdown]
# ## 학습 정리
#
# ### 1. PDF 생성 (reportlab)
# | 요소 | 클래스/함수 |
# |------|------------|
# | 문서 | SimpleDocTemplate |
# | 텍스트 | Paragraph |
# | 테이블 | Table, TableStyle |
# | 이미지 | Image |
# | 간격 | Spacer |
#
# ### 2. Excel 생성 (openpyxl)
# | 작업 | 방법 |
# |------|------|
# | 파일 생성 | Workbook() |
# | DataFrame 변환 | pd.ExcelWriter |
# | 스타일 | Font, PatternFill, Border |
# | 이미지 추가 | ws.add_image() |
#
# ### 3. 리포트 생성 흐름
# 1. 데이터 수집 (pykrx)
# 2. 통계 계산 (`calculate_stock_stats`)
# 3. 차트 이미지 생성 (`create_stock_chart_image`)
# 4. PDF 리포트 생성 (`generate_pdf_report`) - 독립 실행 가능
# 5. Excel 리포트 생성 (`generate_excel_report`) - 독립 실행 가능
#
# **다음 차시**: 38차시 - 일일 주식 리포트 자동 생성 및 이메일 발송
